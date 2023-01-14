#!/usr/bin/env python
# Converts excel sheets to csv.
import os,csv
from pathlib import Path
from openpyxl import load_workbook

excelLocation = Path("chapter_16", "excelSpreadsheets")
if not excelLocation.is_dir():
    print("ERROR: Folder location is missing")
    exit(1)

for file in os.listdir(str(excelLocation)):
    if not file.endswith(".xlsx"):
        continue
    print(f"Working on file: {file}")

    wb = load_workbook(str(Path(excelLocation/file)))
    for ws in wb:
        with open(excelLocation/f"{file.removesuffix('.xlsx')}_{ws.title}.csv",'w',newline="") as outputfile:
            writer = csv.writer(outputfile)
            for x in range(1,ws.max_row+1):
                row = []
                for y in range(1,ws.max_column+1):
                    row.append(ws.cell(row=x, column=y).value)
                writer.writerow(row)
