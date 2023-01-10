#!/usr/bin/env python
import sys
from pathlib import Path
from openpyxl import Workbook

if len(sys.argv) != 2:
    print("Please provide a number")
    exit(1)

try:
    num = int(sys.argv[1])
except ValueError:
    print("Invalid Entry.")
    exit(1)


# create excel document
wb = Workbook()
ws = wb.active
ws.title = f"Multiplication by {num}"
# Populate first column

for ro in range(0,num+1):
    for col in range(0,num+1):
        if col == 0:
            val = ro
        elif ro == 0:
            val = col
        else:
            val = ro * col
        ws.cell(row=ro+1, column=col+1, value=val) # values off set by one

ws["A1"] = ""
wb.save(str(Path("chapter_13/multiplication.xlsx")))
