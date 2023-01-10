#!/usr/bin/env python
import sys
from pathlib import Path
from openpyxl import load_workbook

if len(sys.argv) != 4:
    print("Please provide a row number, number of blank lines and a xlsx file it add blank rows to.")
    exit(1)

try:
    blank_start = int(sys.argv[1])
    blank_count = int(sys.argv[2])
except ValueError:
    print("Invalid Entry for row or number of blank lines.")
    exit(1)

if blank_start <= 0:
    print("Invalid Entry for row, must be greater then or equal to 1.")
    exit(1)

if blank_count <= 0:
    print("Invalid Entry for blank lines, must be greater then or equal to 1.")
    exit(1)

file = sys.argv[3]
if not file.endswith(".xlsx"):
    print("File must end with \".xlsx\".")
    exit(1)

if not Path(file).is_file():
   print("File provided is not a file or does not exist.")
   exit(1)


wb = load_workbook(str(Path(file)))
ws = wb.active
if blank_start > ws.max_row:
    print(f"Row value of {blank_start} is too high, max is {ws.max_row}")

# Rows that will be blanked.
blank_rows = list(range(blank_start, blank_start+blank_count))

buffer = []
for row in range(blank_start,ws.max_row + len(blank_rows) + 1):
    for col in range(1, ws.max_column + 1):
        buffer.append(ws.cell(row=row, column=col).value)
        if row in blank_rows:
            ws.cell(row=row, column=col, value="") # Blank line
        else:
            pop = buffer.pop(0)
            ws.cell(row=row, column=col, value=pop)


wb.save(str(Path(file)))
